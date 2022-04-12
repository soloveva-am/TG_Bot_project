from helpers import get_value_merged, get_color_merged

from typing import Iterator, Tuple
from copy import deepcopy

from openpyxl.worksheet.worksheet import Worksheet

# дни недели
weekdays = (
    "Понедельник",
    "Вторник",
    "Среда",
    "Четверг",
    "Пятница",
    "Суббота",
    "Воскресенье",
)
hour_list = (
    "09:00 – 10:25",
    "10:45 – 12:10",
    "12:20 – 13:45",
    "13:55 – 15:20",
    "15:30 – 16:55",
    "17:05 – 18:30",
    "18:35 – 20:00",
)

# шаблон расписания
blank_timetable = {day: {time: "😴" for time in hour_list} for day in weekdays}

# цветные круги, соответствующие семинарам, лабам / англу, лекциям, базовому дню и военке
colors_to_circles = {
    "#CCFFFF": "🔵",  # семинары
    "#92D050": "🔵",  # семинары
    "#00FFFF": "🔵",  # семинары
    "#66FFFF": "🔵",  # семинары
    "#FFFFFF": "🔵",  # семинары
    "#00B050": "🔵",  # семинары
    "#FFFF99": "🟡",  # лабы / англ
    "#FF99CC": "🔴",  # лекции
    "#CCFFCC": "🟢",  # базовый день
    "#FFC000": "🟠",  # военка
    "#FFCC00": "🟠",  # военка
}


def fix_pair_hours(hours: str):
    hours = hours.split()
    # преобразуем время пары к формату hh:mm – hh:mm
    if len(hours[0][:-2]) == 1:
        hours[0] = "0" + hours[0]
    hours = (
        hours[0][:-2]
        + ":"
        + hours[0][-2:]
        + " – "
        + hours[2][:-2]
        + ":"
        + hours[2][-2:]
    )
    return hours


def get_all_timetable_from_file(table: Worksheet) -> Iterator[Tuple]:
    """
    Функция, которая из таблицы Excel с расписанием выделяет расписание для каждой группы
    и записывает его в базу данных.
    :param table: таблица с расписанием
    :return:
    """
    for j in range(3, table.max_column + 1):  # смотрим на значения по столбцам
        group_name = table.cell(2, j).value  # номер группы
        if group_name in {
            "Дни",
            "Часы",
        }:  # если это не номер группы, то пропускаем столбец
            continue
        # иначе если столбец - это номер группы, то составляем для него расписание
        elif group_name is not None:
            group_name = str(group_name).strip()
            global blank_timetable
            timetable = deepcopy(blank_timetable)
            for k in range(3, table.max_row + 1):  # проходимся по столбцу
                # если клетки относятся ко дню недели (не разделители)
                day, hours, pair = get_value_merged(table, table.cell(k, 1)), None, None
                if day in timetable:
                    hours = get_value_merged(
                        table, table.cell(k, 2)
                    )  # клетка, в которой лежит значение времени
                    pair = get_value_merged(
                        table, table.cell(k, j)
                    )  # клетка, в которой лежит значение пары
                    color = get_color_merged(table, table.cell(k, j))  # цвет клетки

                # рассматриваем только те клетки, для которых определено значение как пары, так и времени
                if hours is not None and pair is not None:
                    hours = fix_pair_hours(hours)
                    # записываем значение в расписание
                    if pair is not None:
                        try:
                            timetable[day][hours] = (
                                colors_to_circles[color] + " " + pair
                            )
                        except KeyError:  # если появится новый цвет, то он будет выведен на экран
                            print(color, pair)
            yield group_name, timetable


# TODO: пересмотреть при загрузке расписания сессии
def get_all_exam_timetable_from_file(table: Worksheet) -> Iterator[Tuple]:
    """
    Функция, которая из таблицы Excel с расписанием экзаменов выделяет расписание для каждой группы
    и записывает его в базу данных.
    :param table: таблица с расписанием
    :return:
    """
    for j in range(3, table.max_column + 1):  # смотрим на значения по столбцам
        group_name = table.cell(7, j).value  # номер группы
        if group_name is not None:
            if isinstance(
                group_name, int
            ):  # если номер группы - просто число, преобразуем его в строку
                group_name = str(group_name)
            # group - словарь с расписанием для группы
            group_name = "".join(group_name.split())
            timetable = dict(Экзамены={})
            for k in range(8, table.max_row + 1):  # проходимся по столбцу
                # если клетки относятся ко дню недели (не разделители)
                date = get_value_merged(table, table.cell(k, 2))  # значение дня
                week_day = get_value_merged(table, table.cell(k, 1))  # день недели
                if date is not None:
                    if date.month == 12:
                        month = "декабря"
                    elif date.month == 1:
                        month = "января"
                    elif date.month == 5:
                        month = "мая"
                    elif date.month == 6:
                        month = "июня"
                    else:
                        month = ""
                    day = (
                        str(date.day) + " " + month + " " + "(" + week_day.lower() + ")"
                    )
                    exam = get_value_merged(
                        table, table.cell(k, j)
                    )  # клетка, в которой лежит значение пары
                    if exam is not None:
                        timetable["Экзамены"][day] = exam
                else:
                    continue
            yield group_name, timetable
